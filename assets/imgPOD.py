import requests
import threading
import os
import datetime

class imgPOD:

    def __init__(self, authToken) -> None:
        self.authToken = authToken
        self.podDir = os.path.join(os.path.expanduser('~'),"Downloads")
        self.data = []
        self.cont = 0
        self.total = 0

    def criarPlanilha(self, data:list):
        import openpyxl
        from openpyxl.drawing.image import Image
        import requests
        from io import BytesIO
        from PIL import Image as PILImage

        def download_and_resize_image(url, width, height):
            response = requests.get(url)
            img = PILImage.open(BytesIO(response.content))
            img = img.resize((width, height), PILImage.LANCZOS)  # Use LANCZOS para redimensionar
            byte_arr = BytesIO()
            img.save(byte_arr, format='PNG')
            byte_arr.seek(0)
            return Image(byte_arr)

        def create_excel_with_images(data, excel_file):
            # Cria uma nova planilha
            wb = openpyxl.Workbook()
            ws = wb.active

            # Cabeçalhos
            ws.cell(row=1, column=1, value="ID JMS")
            ws.cell(row=1, column=2, value="LINK IMAGEM")
            ws.cell(row=1, column=3, value="IMAGEM POD")

            # Define o tamanho das células (em pixels)
            cell_width = 100  # Largura da célula em pixels
            cell_height = 100  # Altura da célula em pixels

            # Converte o tamanho das células para o formato do Excel
            excel_cell_width = cell_width * 0.14
            excel_cell_height = cell_height * 0.75

            # Insere os dados
            row = 2
            for item in data:
                id_value = item[0]
                image_links = item[1]

                # Insere o ID
                ws.cell(row=row, column=1, value=id_value)

                # Insere os links de imagens em células separadas
                col = 2
                for url in image_links:
                    # Insere o link na célula
                    ws.cell(row=row, column=col, value=url)
                    col += 1

                # Insere cada imagem em colunas adjacentes ao último link
                for idx, url in enumerate(image_links):
                    try:
                        # Baixa e redimensiona a imagem
                        img = download_and_resize_image(url, cell_width, cell_height)
                        
                        # A coluna para a imagem será a próxima coluna livre após os links
                        img_col = col + idx
                        cell_position = f'{chr(64 + img_col)}{row}'
                        
                        # Insere a imagem na célula
                        ws.add_image(img, cell_position)
                        
                        # Ajusta a largura da coluna e a altura da linha
                        ws.column_dimensions[chr(64 + img_col)].width = excel_cell_width
                        ws.row_dimensions[row].height = excel_cell_height
                        
                    except Exception as e:
                        print(f"Erro ao baixar ou inserir a imagem de {url}: {e} {item[0]}")

                row += 1

            # Salva a planilha
            wb.save(excel_file)
        # Nome do arquivo Excel a ser salvo
        excel_file = os.path.join(self.podDir, f"POD - {datetime.datetime.today().strftime('%d-%m-%Y %H-%M-%S')}.xlsx")

        # Chama a função para criar a planilha com as imagens
        create_excel_with_images(data, excel_file)

    def retrieveScanByCode(self, jmsId):

        #POST
        url = "https://gw.jtjms-br.com/operatingplatform/podTracking/inner/query/keywordList"
        header = {
            'Accept' : 'application/json, text/plain, */*',
            'Accept-Encoding' : 'gzip, deflate, br, zstd',
            'Accept-Language' : 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7,de;q=0.6',
            'Cache-Control' : 'max-age=2, must-revalidate',
            'Connection' : 'keep-alive',
            'Content-Length' : '80',
            'Content-Type' : 'application/json;charset=UTF-8',
            'Host' : 'gw.jtjms-br.com',
            'Origin' : 'https',
            'Referer' : 'https',
            'Sec-Fetch-Dest' : 'empty',
            'Sec-Fetch-Mode' : 'cors',
            'Sec-Fetch-Site' : 'same-site',
            'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'authToken' : self.authToken,
            'lang' : 'PT',
            'langType' : 'PT',
            'routeName' : 'trackingExpress',
            'sec-ch-ua' : '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile' : '?0',
            'sec-ch-ua-platform' : '"Windows"',
            'timezone' : 'GMT-0300'
            }
        payload = {
            "keywordList":[jmsId],
            "trackingTypeEnum":"WAYBILL",
            "countryId":"1"
            }

        result = requests.post(url = url, headers = header, json = payload)
        if result.status_code == 200:
            results = result.json().get("data",[])[0].get("details",[])
            for result in results:
                if result["scanTypeName"] == "assinatura de encomenda":
                    return (result["billCode"], result["scanTime"], result["scanByCode"], result["imgType"])

    def retrieveImg(self, idSublist, widget, widget2):

        for id in idSublist:
            self.cont += 1
            info = self.retrieveScanByCode(id)
            if info:
                # POST
                url = "https://gw.jtjms-br.com/operatingplatform/podTracking/img/path"
                header = {
                    'Accept' : 'application/json, text/plain, */*',
                    'Accept-Encoding' : 'gzip, deflate, br, zstd',
                    'Accept-Language' : 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7,de;q=0.6',
                    'Cache-Control' : 'max-age=2, must-revalidate',
                    'Connection' : 'keep-alive',
                    'Content-Length' : '119',
                    'Content-Type' : 'application/json;charset=UTF-8',
                    'Host' : 'gw.jtjms-br.com',
                    'Origin' : 'https',
                    'Referer' : 'https',
                    'Sec-Fetch-Dest' : 'empty',
                    'Sec-Fetch-Mode' : 'cors',
                    'Sec-Fetch-Site' : 'same-site',
                    'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
                    'authToken' : self.authToken,
                    'lang' : 'PT',
                    'langType' : 'PT',
                    'routeName' : 'trackingExpress',
                    'sec-ch-ua' : '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
                    'sec-ch-ua-mobile' : '?0',
                    'sec-ch-ua-platform' : '"Windows"',
                    'timezone' : 'GMT-0300'
                    }
                payload = {"waybillNo":info[0],"scanTime":info[1],"scanByCode":info[2],"imgType":info[3],"countryId":"1"}

                result = requests.post(url = url, headers = header, json = payload)
                if result.status_code == 200:
                    links = []
                    for link in result.json().get("data",[]):
                        links.append(link)
                    self.data.append([info[0], links])

                porcentagem = (self.cont/self.total) * 100
                widget.configure(text = f"{round(porcentagem,2)}%")
                widget2.configure(value = self.cont, maximum = self.total)

            else:
                self.data.append([id, []])
        widget.configure(text = f"Aguarde... Gerando planilha.")

    def start(self, idList:list, widget, widget2):

        size = len(idList)
        global total
        total = 0
        total = size
        self.total = total
        threads = []
        step = max(round(size * 0.025), 1)
        sublists = [idList[i:i + step] for i in range(0, size, step)]

        for sublist in sublists:
            thread = threading.Thread(target=self.retrieveImg, args=(sublist, widget, widget2))
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        self.criarPlanilha(self.data)


# imgPOD("1aa9e9ac08ee4458ba1bab626a68b513").start(["888030073614848","888000546099260"])